from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from urllib.parse import parse_qs, urlparse
from email.parser import BytesParser
from email.policy import default
from pathlib import Path
from datetime import datetime, timezone
import json
import mimetypes
import os
import secrets
import smtplib
import sqlite3
import sys

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
STATIC_DIR = ROOT / "static"
DATA_DIR = Path(os.environ.get("QUOTE_DATA_DIR", ROOT / "data"))
UPLOAD_DIR = DATA_DIR / "uploads"
DB_PATH = DATA_DIR / "quotes.db"


def now_iso():
    return datetime.now(timezone.utc).isoformat()


def ensure_dirs():
    DATA_DIR.mkdir(exist_ok=True)
    UPLOAD_DIR.mkdir(exist_ok=True)


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    ensure_dirs()
    with db() as conn:
        conn.executescript(
            """
            create table if not exists requests (
              id integer primary key autoincrement,
              token text not null unique,
              project_name text not null,
              due_date text,
              memo text,
              uploaded_file text,
              created_at text not null
            );

            create table if not exists parts (
              id integer primary key autoincrement,
              request_id integer not null references requests(id),
              category text,
              name text not null,
              spec text,
              quantity integer not null default 1,
              baseline_unit_price real,
              unit text,
              note text
            );

            create table if not exists vendors (
              id integer primary key autoincrement,
              request_id integer not null references requests(id),
              token text not null unique,
              company_name text not null,
              contact_name text,
              contact_email text,
              status text not null default 'pending',
              submitted_at text
            );

            create table if not exists quote_lines (
              id integer primary key autoincrement,
              vendor_id integer not null references vendors(id),
              part_id integer not null references parts(id),
              unit_price real,
              lead_time text,
              maker text,
              model text,
              note text
            );

            create table if not exists notifications (
              id integer primary key autoincrement,
              request_id integer not null references requests(id),
              vendor_id integer references vendors(id),
              channel text not null,
              recipient text,
              subject text,
              body text,
              status text not null,
              created_at text not null
            );
            """
        )
        part_columns = {
            row["name"]
            for row in conn.execute("pragma table_info(parts)")
        }
        if "baseline_unit_price" not in part_columns:
            conn.execute("alter table parts add column baseline_unit_price real")


def row_dict(row):
    return dict(row) if row else None


def parse_excel(path):
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 파일에 데이터가 없습니다.")

    header_idx = None
    headers = []
    for idx, row in enumerate(rows[:10]):
        normalized = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if any(value in normalized for value in ["품목", "부품명", "name", "part", "item"]):
            header_idx = idx
            headers = normalized
            break

    if header_idx is None:
        header_idx = 0
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]

    def find_col(candidates):
        for candidate in candidates:
            candidate = candidate.lower()
            for idx, header in enumerate(headers):
                if candidate == header or candidate in header:
                    return idx
        return None

    col_category = find_col(["구분", "분류", "category", "type"])
    col_name = find_col(["품목", "부품명", "제품명", "name", "part", "item"])
    col_spec = find_col(["사양", "규격", "spec", "description"])
    col_qty = find_col(["수량", "qty", "quantity"])
    col_unit_price = find_col(["단가", "unit price", "unit_price", "price"])
    col_unit = find_col(["단위", "unit"])
    col_note = find_col(["비고", "메모", "note", "remark"])

    if col_name is None:
        col_name = 0

    parts = []
    for row in rows[header_idx + 1 :]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        def value(col):
            if col is None or col >= len(row) or row[col] is None:
                return ""
            return str(row[col]).strip()

        name = value(col_name)
        spec = value(col_spec)
        if not name and not spec:
            continue

        qty_raw = value(col_qty)
        try:
            quantity = int(float(qty_raw)) if qty_raw else 1
        except ValueError:
            quantity = 1

        unit_price_raw = value(col_unit_price).replace(",", "")
        try:
            baseline_unit_price = float(unit_price_raw) if unit_price_raw else None
        except ValueError:
            baseline_unit_price = None

        parts.append(
            {
                "category": value(col_category),
                "name": name or spec,
                "spec": spec,
                "quantity": max(quantity, 1),
                "baseline_unit_price": baseline_unit_price,
                "unit": value(col_unit) or "EA",
                "note": value(col_note),
            }
        )

    if not parts:
        raise ValueError("부품 행을 찾지 못했습니다. '품목/부품명, 사양, 수량' 형태의 헤더가 필요합니다.")
    return parts


def send_notification(request_row, vendor_row):
    to_email = os.environ.get("QUOTE_NOTIFY_EMAIL", "").strip()
    subject = f"[견적 제출] {request_row['project_name']} - {vendor_row['company_name']}"
    body = (
        f"사업명: {request_row['project_name']}\n"
        f"업체명: {vendor_row['company_name']}\n"
        f"제출 시간: {vendor_row['submitted_at']}\n"
        f"관리 화면에서 견적을 확인하세요.\n"
    )

    status = "logged"
    if to_email and os.environ.get("SMTP_HOST"):
        try:
            host = os.environ["SMTP_HOST"]
            port = int(os.environ.get("SMTP_PORT", "587"))
            user = os.environ.get("SMTP_USER", "")
            password = os.environ.get("SMTP_PASSWORD", "")
            sender = os.environ.get("SMTP_FROM", user or to_email)
            with smtplib.SMTP(host, port, timeout=10) as smtp:
                smtp.starttls()
                if user:
                    smtp.login(user, password)
                message = (
                    f"From: {sender}\r\n"
                    f"To: {to_email}\r\n"
                    f"Subject: {subject}\r\n"
                    f"Content-Type: text/plain; charset=utf-8\r\n\r\n"
                    f"{body}"
                )
                smtp.sendmail(sender, [to_email], message.encode("utf-8"))
            status = "sent"
        except Exception as exc:
            status = f"failed: {exc}"

    with db() as conn:
        conn.execute(
            """
            insert into notifications
              (request_id, vendor_id, channel, recipient, subject, body, status, created_at)
            values (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                request_row["id"],
                vendor_row["id"],
                "email",
                to_email or "not configured",
                subject,
                body,
                status,
                now_iso(),
            ),
        )


class AppHandler(BaseHTTPRequestHandler):
    server_version = "QuotePortal/0.1"

    def log_message(self, fmt, *args):
        return

    def send_json(self, data, status=200):
        payload = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def send_text(self, text, status=200):
        payload = text.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def read_json(self):
        length = int(self.headers.get("Content-Length", 0))
        if length == 0:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/"):
            return self.handle_api_get(parsed)
        return self.serve_static(parsed.path)

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/"):
            return self.handle_api_post(parsed)
        return self.send_text("Not found", 404)

    def serve_static(self, path):
        if path == "/":
            path = "/index.html"
        target = (STATIC_DIR / path.lstrip("/")).resolve()
        if not str(target).startswith(str(STATIC_DIR.resolve())) or not target.exists():
            target = STATIC_DIR / "index.html"
        content = target.read_bytes()
        content_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        if target.suffix == ".html":
            content_type = "text/html; charset=utf-8"
        elif target.suffix == ".css":
            content_type = "text/css; charset=utf-8"
        elif target.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def handle_api_get(self, parsed):
        path = parsed.path
        query = parse_qs(parsed.query)

        if path == "/api/requests":
            with db() as conn:
                requests = [
                    dict(row)
                    for row in conn.execute(
                        """
                        select r.*,
                          (select count(*) from vendors v where v.request_id = r.id) as vendor_count,
                          (select count(*) from vendors v where v.request_id = r.id and v.status = 'submitted') as submitted_count
                        from requests r
                        order by r.created_at desc
                        """
                    )
                ]
            return self.send_json({"requests": requests})

        if path == "/api/request":
            token = query.get("token", [""])[0]
            with db() as conn:
                request = row_dict(conn.execute("select * from requests where token = ?", (token,)).fetchone())
                if not request:
                    return self.send_json({"error": "요청을 찾을 수 없습니다."}, 404)
                parts = [dict(row) for row in conn.execute("select * from parts where request_id = ?", (request["id"],))]
                vendors = [
                    dict(row)
                    for row in conn.execute(
                        "select * from vendors where request_id = ? order by company_name", (request["id"],)
                    )
                ]
                quote_rows = [
                    dict(row)
                    for row in conn.execute(
                        """
                        select q.*, p.name as part_name, p.quantity, v.company_name
                        from quote_lines q
                        join parts p on p.id = q.part_id
                        join vendors v on v.id = q.vendor_id
                        where p.request_id = ?
                        order by p.id, v.company_name
                        """,
                        (request["id"],),
                    )
                ]
                notifications = [
                    dict(row)
                    for row in conn.execute(
                        "select * from notifications where request_id = ? order by created_at desc limit 20",
                        (request["id"],),
                    )
                ]
            return self.send_json(
                {
                    "request": request,
                    "parts": parts,
                    "vendors": vendors,
                    "quotes": quote_rows,
                    "notifications": notifications,
                }
            )

        if path == "/api/vendor":
            token = query.get("token", [""])[0]
            with db() as conn:
                vendor = row_dict(conn.execute("select * from vendors where token = ?", (token,)).fetchone())
                if not vendor:
                    return self.send_json({"error": "견적 링크를 찾을 수 없습니다."}, 404)
                request = row_dict(conn.execute("select * from requests where id = ?", (vendor["request_id"],)).fetchone())
                parts = [dict(row) for row in conn.execute("select * from parts where request_id = ?", (request["id"],))]
                lines = [
                    dict(row)
                    for row in conn.execute("select * from quote_lines where vendor_id = ?", (vendor["id"],))
                ]
            return self.send_json({"request": request, "vendor": vendor, "parts": parts, "lines": lines})

        return self.send_json({"error": "Not found"}, 404)

    def handle_api_post(self, parsed):
        if parsed.path == "/api/requests":
            return self.create_request()
        if parsed.path == "/api/quote":
            return self.submit_quote()
        return self.send_json({"error": "Not found"}, 404)

    def create_request(self):
        content_type = self.headers.get("Content-Type", "")
        if "multipart/form-data" not in content_type:
            return self.send_json({"error": "Excel 파일을 포함한 폼으로 보내주세요."}, 400)

        length = int(self.headers.get("Content-Length", 0))
        message = BytesParser(policy=default).parsebytes(
            f"Content-Type: {content_type}\r\n\r\n".encode("utf-8") + self.rfile.read(length)
        )

        def field(name, default_value=""):
            part = message.get_payload()
            for item in part:
                if item.get_param("name", header="content-disposition") == name:
                    return item.get_content().strip()
            return default_value

        project_name = field("projectName")
        due_date = field("dueDate")
        memo = field("memo")
        vendors_raw = field("vendors")
        manual_parts_raw = field("manualParts", "[]")
        if not project_name:
            return self.send_json({"error": "사업명을 입력해주세요."}, 400)

        file_part = None
        for item in message.get_payload():
            if item.get_param("name", header="content-disposition") == "file":
                file_part = item
                break
        saved_name = ""
        parts = []
        if file_part is not None and file_part.get_filename():
            filename = Path(file_part.get_filename()).name
            saved_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{secrets.token_hex(4)}_{filename}"
            saved_path = UPLOAD_DIR / saved_name
            saved_path.write_bytes(file_part.get_payload(decode=True))

            try:
                parts = parse_excel(saved_path)
            except Exception as exc:
                return self.send_json({"error": str(exc)}, 400)
        else:
            try:
                manual_parts = json.loads(manual_parts_raw)
            except json.JSONDecodeError:
                manual_parts = []
            for part in manual_parts:
                name = str(part.get("name", "")).strip()
                if not name:
                    continue
                quantity_raw = str(part.get("quantity", "1")).replace(",", "").strip()
                unit_price_raw = str(part.get("unitPrice", "")).replace(",", "").strip()
                try:
                    quantity = int(float(quantity_raw)) if quantity_raw else 1
                except ValueError:
                    quantity = 1
                try:
                    unit_price = float(unit_price_raw) if unit_price_raw else None
                except ValueError:
                    unit_price = None
                parts.append(
                    {
                        "category": "",
                        "name": name,
                        "spec": "",
                        "quantity": max(quantity, 1),
                        "baseline_unit_price": unit_price,
                        "unit": "EA",
                        "note": "",
                    }
                )
            if not parts:
                return self.send_json({"error": "Excel 파일을 올리거나 수기 부품을 1개 이상 입력해주세요."}, 400)

        vendors = []
        for line in vendors_raw.splitlines():
            clean = line.strip()
            if not clean:
                continue
            chunks = [chunk.strip() for chunk in clean.split(",")]
            vendors.append(
                {
                    "company_name": chunks[0],
                    "contact_name": chunks[1] if len(chunks) > 1 else "",
                    "contact_email": chunks[2] if len(chunks) > 2 else "",
                }
            )
        if not vendors:
            return self.send_json({"error": "최소 1개 업체를 입력해주세요."}, 400)

        request_token = secrets.token_urlsafe(18)
        with db() as conn:
            cur = conn.execute(
                """
                insert into requests (token, project_name, due_date, memo, uploaded_file, created_at)
                values (?, ?, ?, ?, ?, ?)
                """,
                (request_token, project_name, due_date, memo, saved_name, now_iso()),
            )
            request_id = cur.lastrowid
            for part in parts:
                conn.execute(
                    """
                    insert into parts
                      (request_id, category, name, spec, quantity, baseline_unit_price, unit, note)
                    values (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        request_id,
                        part["category"],
                        part["name"],
                        part["spec"],
                        part["quantity"],
                        part.get("baseline_unit_price"),
                        part["unit"],
                        part["note"],
                    ),
                )
            vendor_links = []
            for vendor in vendors:
                vendor_token = secrets.token_urlsafe(24)
                conn.execute(
                    """
                    insert into vendors
                      (request_id, token, company_name, contact_name, contact_email)
                    values (?, ?, ?, ?, ?)
                    """,
                    (
                        request_id,
                        vendor_token,
                        vendor["company_name"],
                        vendor["contact_name"],
                        vendor["contact_email"],
                    ),
                )
                vendor_links.append({**vendor, "token": vendor_token})

        base_url = f"http://{self.headers.get('Host', 'localhost:8000')}"
        return self.send_json(
            {
                "requestToken": request_token,
                "adminUrl": f"{base_url}/admin.html?token={request_token}",
                "vendorLinks": [
                    {
                        **vendor,
                        "url": f"{base_url}/vendor.html?token={vendor['token']}",
                    }
                    for vendor in vendor_links
                ],
                "partCount": len(parts),
            }
        )

    def submit_quote(self):
        payload = self.read_json()
        token = payload.get("token", "")
        lines = payload.get("lines", [])
        has_price = False
        with db() as conn:
            vendor = row_dict(conn.execute("select * from vendors where token = ?", (token,)).fetchone())
            if not vendor:
                return self.send_json({"error": "견적 링크를 찾을 수 없습니다."}, 404)
            was_submitted = vendor["status"] == "submitted"
            request = row_dict(conn.execute("select * from requests where id = ?", (vendor["request_id"],)).fetchone())
            part_ids = {
                row["id"]
                for row in conn.execute("select id from parts where request_id = ?", (vendor["request_id"],))
            }
            conn.execute("delete from quote_lines where vendor_id = ?", (vendor["id"],))
            for line in lines:
                part_id = int(line.get("partId", 0))
                if part_id not in part_ids:
                    continue
                unit_price_raw = str(line.get("unitPrice", "")).replace(",", "").strip()
                try:
                    unit_price = float(unit_price_raw) if unit_price_raw else None
                except ValueError:
                    unit_price = None
                if unit_price is not None:
                    has_price = True
                conn.execute(
                    """
                    insert into quote_lines
                      (vendor_id, part_id, unit_price, lead_time, maker, model, note)
                    values (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        vendor["id"],
                        part_id,
                        unit_price,
                        line.get("leadTime", ""),
                        line.get("maker", ""),
                        line.get("model", ""),
                        line.get("note", ""),
                    ),
                )
            submitted_at = now_iso()
            if has_price:
                conn.execute(
                    "update vendors set status = 'submitted', submitted_at = ? where id = ?",
                    (submitted_at, vendor["id"]),
                )
            vendor = row_dict(conn.execute("select * from vendors where id = ?", (vendor["id"],)).fetchone())

        if has_price and not was_submitted:
            send_notification(request, vendor)
        return self.send_json({"ok": True, "submittedAt": submitted_at})


def main():
    init_db()
    port = int(os.environ.get("PORT", "8000"))
    host = os.environ.get("HOST", "127.0.0.1")
    if os.environ.get("RENDER"):
        host = "0.0.0.0"
    server = ThreadingHTTPServer((host, port), AppHandler)
    print(f"Quote portal running at http://{host}:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped")


if __name__ == "__main__":
    main()
